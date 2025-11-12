import os
import re
import sys
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import tabula
import pdfplumber
import tkinter as tk
from tkinter import messagebox
import pandas as pd


def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def setup_logging(base_dir):
    logs_dir = os.path.join(base_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)
    log_path = os.path.join(logs_dir, "extract_log.log")
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        filemode="w"  # overwrite each run
    )
    return log_path


def load_config(base_dir):
    config_path = os.path.join(base_dir, "config.json")
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path, "r") as f:
        return json.load(f)


def generate_sheet_name(pdf_name):
    sheet_name = re.sub(r"\.pdf$", "", os.path.basename(pdf_name), flags=re.IGNORECASE)
    sheet_name = re.sub(r"[\\/:?*\[\]]", "_", sheet_name)
    return sheet_name[:31]


def extract_metadata(pdf_path):
    metadata = {
        "AUDIT ID": "",
        "NCPDP": "",
        "Date": "",
        "Address": "",
        "Subject": ""
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None
            full_text = "\n".join([page.extract_text() or "" for page in pdf.pages])

        if not full_text.strip():
            return None

        audit_match = re.search(r"AUDIT\s*ID[:\s]*([A-Za-z0-9\-]+)", full_text, re.IGNORECASE)
        ncpdp_match = re.search(r"NCPDP[:\s]*([0-9]+)", full_text, re.IGNORECASE)
        date_match = re.search(r"Date[:\s]*(\d{1,2}/\d{1,2}/\d{2,4})", full_text, re.IGNORECASE)

        metadata["AUDIT ID"] = audit_match.group(1).strip() if audit_match else ""
        metadata["NCPDP"] = ncpdp_match.group(1).strip() if ncpdp_match else ""
        metadata["Date"] = date_match.group(1).strip() if date_match else ""

        addr_match = re.search(
            r"(?m)^(?:[A-Z\s]*\bPHARMACY\b[^\n]*\n(?:.*\n){0,5}?FAX[:\s]*\d+)",
            full_text, re.IGNORECASE
        )
        if addr_match:
            addr_text = addr_match.group(0)
            addr_text = re.sub(r"\s*\n\s*", ", ", addr_text)
            addr_text = re.sub(r"\s{2,}", " ", addr_text).strip(" ,")
            metadata["Address"] = addr_text

        subject_match = re.search(r"RE[:\s].{5,}", full_text)
        if subject_match:
            metadata["Subject"] = subject_match.group(0).strip()

        return metadata
    except Exception as e:
        logging.error(f"Metadata extraction failed for {pdf_path}: {e}")
        return metadata


def extract_tables(pdf_path):
    """Extract clean tables with proper headers and no duplicates."""
    try:
        all_tables = []
        seen_titles = set()

        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                page_tables = tabula.read_pdf(pdf_path, pages=page_num, multiple_tables=True, lattice=True)
                if not page_tables:
                    continue

                detected_tables = page.find_tables()
                words = page.extract_words()
                line_map = {}
                for w in words:
                    y_key = round(w["bottom"], 1)
                    if y_key not in line_map:
                        line_map[y_key] = []
                    line_map[y_key].append(w["text"])
                line_texts = {k: " ".join(v).strip() for k, v in line_map.items()}

                for idx, table in enumerate(page_tables):
                    if not isinstance(table, pd.DataFrame) or table.empty:
                        continue

                    # Remove rows that are completely empty
                    table = table.dropna(how='all')
                    table = table.reset_index(drop=True)

                    title = ""
                    try:
                        if detected_tables and len(detected_tables) > idx:
                            x0, top, x1, bottom = detected_tables[idx].bbox
                            above_lines = {
                                y: text for y, text in line_texts.items() if y < top and (top - y) < 120
                            }
                            if above_lines:
                                nearest_y = max(above_lines.keys())
                                title = above_lines[nearest_y].strip()
                    except Exception as te:
                        logging.warning(f"Could not extract table title on page {page_num}: {te}")

                    if not title:
                        # Try to detect header-like first row text
                        possible_title = str(table.columns[0])
                        if re.search(r"[A-Za-z]", possible_title):
                            title = possible_title
                        else:
                            title = f"Table_{page_num}_{idx+1}"

                    if title in seen_titles:
                        continue  # avoid duplicate tables

                    seen_titles.add(title)
                    all_tables.append({"title": title, "data": table})

        logging.info(f"Extracted {len(all_tables)} unique tables from {os.path.basename(pdf_path)}")
        return all_tables
    except Exception as e:
        logging.error(f"Table extraction failed for {pdf_path}: {e}")
        return []


def write_to_excel(pdf_data, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for pdf_file, content in pdf_data.items():
        ws = wb.create_sheet(title=generate_sheet_name(pdf_file))
        current_row = 1

        metadata = content.get("metadata")
        tables = content.get("tables", [])

        if metadata is None:
            ws.cell(row=current_row, column=1, value=f"No content in PDF {os.path.basename(pdf_file)}")
            continue

        current_row += 1
        for field, value in metadata.items():
            ws.cell(row=current_row, column=1, value=field)
            ws.cell(row=current_row, column=2, value=value)
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=2).border = border_style
            current_row += 1
        current_row += 2

        if not tables:
            ws.cell(row=current_row, column=1, value="No tables found in this PDF.")
            continue

        for table_info in tables:
            table = table_info["data"]
            title = table_info["title"]

            ws.cell(row=current_row, column=1, value="Table:").font = bold
            ws.cell(row=current_row, column=2, value=title)
            current_row += 1  # immediate next line for headers

            # Write headers
            for col_idx, col_name in enumerate(table.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
                cell.font = bold
                cell.border = border_style
            current_row += 1

            # Write data rows (skip extra empty rows)
            for _, row in table.iterrows():
                if not any(pd.notna(v) and str(v).strip() != "" for v in row):
                    continue  # skip completely empty rows
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx,
                                   value=str(value) if pd.notna(value) else "")
                    cell.border = border_style
                current_row += 1

            current_row += 2  # spacing between tables

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 15)

    # ✅ Overwrite Excel file (same name every run)
    output_path = os.path.join(os.path.dirname(output_path), "extracted_table_data.xlsx")
    wb.save(output_path)
    logging.info(f"✅ Excel file saved: {output_path}")


def show_message(title, message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()


def main():
    base_dir = get_base_dir()
    log_path = setup_logging(base_dir)
    logging.info("Starting PDF Table Extractor...")

    try:
        config = load_config(base_dir)
    except Exception as e:
        msg = f"Failed to load config.json: {e}"
        logging.error(msg)
        show_message("Error", msg)
        return

    input_path = config.get("input_path")
    output_dir = config.get("output_dir")

    if not input_path or not os.path.exists(input_path):
        msg = f"Invalid input path: {input_path}"
        logging.error(msg)
        show_message("Error", msg)
        return

    os.makedirs(output_dir, exist_ok=True)
    output_excel = os.path.join(output_dir, "extracted_table_data.xlsx")

    if os.path.isdir(input_path):
        pdf_files = [os.path.join(input_path, f)
                     for f in os.listdir(input_path) if f.lower().endswith(".pdf")]
    elif input_path.lower().endswith(".pdf"):
        pdf_files = [input_path]
    else:
        pdf_files = []

    if not pdf_files:
        msg = f"No PDF files found in {input_path}"
        logging.warning(msg)
        show_message("Warning", msg)
        return

    pdf_data = {}
    for pdf_file in pdf_files:
        logging.info(f"Processing: {pdf_file}")
        metadata = extract_metadata(pdf_file)
        tables = extract_tables(pdf_file)
        pdf_data[pdf_file] = {"metadata": metadata, "tables": tables}

    if pdf_data:
        write_to_excel(pdf_data, output_excel)
        show_message("Success", f"Extraction completed.\nExcel saved to:\n{output_excel}")
    else:
        msg = "No data extracted from any files."
        logging.warning(msg)
        show_message("Warning", msg)

    logging.info(f"Log file: {log_path}")


if __name__ == "__main__":
    main()
